from openpyxl import Workbook
from openpyxl.styles import Font
from config import OUTPUT_FILE


def save_jobs(jobs):
    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    ws.append([
        "Company",
        "Position",
        "Location",
        "Apply",
        "Overall Score",
        "Engineering",
        "Technical",
        "Sales",
        "Customer",
        "Telecom",
        "Remote",
        "Matched Skills"
    ])

    for job in jobs:
        apply_link = job.get("Apply", "")

        ws.append([
            job.get("Company", ""),
            job.get("Position", ""),
            job.get("Location", ""),
            apply_link,
            job.get("Score", 0),
            job.get("Engineering Score", 0),
            job.get("Technical Score", 0),
            job.get("Sales Score", 0),
            job.get("Customer Score", 0),
            job.get("Telecom Score", 0),
            job.get("Remote Score", 0),
            job.get("Matched Skills", "")
        ])

        # Make Apply link clickable
        cell = ws.cell(row=ws.max_row, column=4)

        if apply_link:
            cell.hyperlink = apply_link
            cell.style = "Hyperlink"

    # Make header bold
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT_FILE)

    print(f"Saved {len(jobs)} jobs to {OUTPUT_FILE}")